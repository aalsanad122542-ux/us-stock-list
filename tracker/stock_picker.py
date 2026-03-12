"""
stock_picker.py
Picks N stocks sequentially from the Excel list to ensure all items are scanned eventually.
"""

import os
import json
import logging
import openpyxl

logger = logging.getLogger(__name__)

_EXCEL_PATH = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "us_listed_stocks.xlsx")
_STATE_FILE = os.environ.get("EMAIL_STATE_FILE", "/tmp/email_state.json")

def _load_state() -> dict:
    if os.path.exists(_STATE_FILE):
        try:
            with open(_STATE_FILE, 'r') as f:
                return json.load(f)
        except Exception as e:
            logger.error(f"Failed to read state file: {e}")
    return {"last_row": 1} # Header is row 1

def _save_state(state: dict):
    os.makedirs(os.path.dirname(_STATE_FILE), exist_ok=True)
    try:
        with open(_STATE_FILE, 'w') as f:
            json.dump(state, f)
    except Exception as e:
        logger.error(f"Failed to save state file: {e}")

def get_next_n_stocks(n: int = 5) -> list[dict]:
    """
    Reads the next N rows from the Excel file, wrapping around to the beginning if needed.
    Returns a list of dicts with ticker and company information.
    """
    if not os.path.exists(_EXCEL_PATH):
        logger.error(f"Excel file not found at {_EXCEL_PATH}")
        return []
        
    state = _load_state()
    last_row = state.get("last_row", 1)
    
    try:
        logger.info(f"Loading '{_EXCEL_PATH}'...")
        wb = openpyxl.load_workbook(_EXCEL_PATH, data_only=True)
        sheet = wb.active
        max_row = sheet.max_row
        
        # If the file is empty or only has headers
        if max_row <= 1:
            logger.warning("Excel file is empty or only contains headers.")
            return []
            
        stocks = []
        current_row = last_row + 1
        
        # Read N rows
        while len(stocks) < n:
            if current_row > max_row:
                current_row = 2 # Wrap around to beginning (skip header)
                
            ticker_cell = sheet.cell(row=current_row, column=1)
            company_cell = sheet.cell(row=current_row, column=2)
            
            ticker = str(ticker_cell.value).strip() if ticker_cell.value else ""
            company = str(company_cell.value).strip() if company_cell.value else ""
            
            if ticker:
                stocks.append({
                    "ticker": ticker,
                    "company": company,
                    "row": current_row
                })
            
            # If we've looped entirely around and found nothing, break to avoid infinite loop
            if current_row == last_row and not ticker:
                 break
                 
            current_row += 1
            
        # Update state with the row of the last stock we picked
        if stocks:
            state["last_row"] = stocks[-1]["row"]
            _save_state(state)
            
        return stocks

    except Exception as e:
        logger.error(f"Error reading Excel file: {e}")
        return []

if __name__ == "__main__":
    logging.basicConfig(level=logging.INFO)
    stocks = get_next_n_stocks(5)
    print(json.dumps(stocks, indent=2))
